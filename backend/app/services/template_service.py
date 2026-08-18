from __future__ import annotations

import csv
import io
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.import_template import (
    ImportTemplateDetail,
    ImportTemplateListResponse,
    ImportTemplateSummary,
    TemplateColumnSchema,
)
from app.templates.definitions import ImportTemplateDefinition, get_template, list_templates

Domain = Literal["expense", "revenue", "fund", "dashboard"]


class TemplateNotFoundError(Exception):
    def __init__(self, code: str) -> None:
        self.code = code
        super().__init__(f"Import template not found: {code}")


class TemplateService:
    def list_templates(self, domain: Domain | None = None) -> ImportTemplateListResponse:
        templates = list_templates(domain)
        items = [self._to_summary(t) for t in templates]
        return ImportTemplateListResponse(items=items, total=len(items))

    def get_template_detail(self, code: str, api_prefix: str) -> ImportTemplateDetail:
        template = self._require_template(code)
        summary = self._to_summary(template)
        return ImportTemplateDetail(
            **summary.model_dump(),
            columns=[self._to_column_schema(col) for col in template.columns],
            download_csv_url=f"{api_prefix}/import/templates/{code}/download",
            download_xlsx_url=f"{api_prefix}/import/templates/{code}/download.xlsx",
        )

    def build_csv(self, code: str) -> tuple[str, bytes]:
        template = self._require_template(code)
        buffer = io.StringIO()
        writer = csv.writer(buffer, lineterminator="\n")
        writer.writerow([col.key for col in template.columns])
        writer.writerow([col.label for col in template.columns])
        for col in template.columns:
            writer.writerow([col.example])
        content = buffer.getvalue().encode("utf-8-sig")
        filename = f"{code}_v{template.version}.csv"
        return filename, content

    def build_xlsx(self, code: str) -> tuple[str, bytes]:
        template = self._require_template(code)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "data"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        label_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")

        for index, col in enumerate(template.columns, start=1):
            key_cell = sheet.cell(row=1, column=index, value=col.key)
            key_cell.font = header_font
            key_cell.fill = header_fill

            label_cell = sheet.cell(row=2, column=index, value=col.label)
            label_cell.font = Font(bold=True)
            label_cell.fill = label_fill

            sheet.cell(row=3, column=index, value=col.example)
            sheet.column_dimensions[get_column_letter(index)].width = max(len(col.label), len(col.key)) + 4

        dict_sheet = workbook.create_sheet("字段说明")
        dict_sheet.append(["column_key", "中文名称", "必填", "类型", "说明", "参考取值", "示例"])
        for col in template.columns:
            dict_sheet.append([
                col.key,
                col.label,
                "是" if col.required else "否",
                col.column_type,
                col.description,
                " / ".join(col.enum_values),
                col.example,
            ])

        output = io.BytesIO()
        workbook.save(output)
        filename = f"{code}_v{template.version}.xlsx"
        return filename, output.getvalue()

    def _require_template(self, code: str) -> ImportTemplateDefinition:
        template = get_template(code)
        if template is None:
            raise TemplateNotFoundError(code)
        return template

    def _to_summary(self, template: ImportTemplateDefinition) -> ImportTemplateSummary:
        required_count = sum(1 for col in template.columns if col.required)
        return ImportTemplateSummary(
            code=template.code,
            name=template.name,
            domain=template.domain,
            version=template.version,
            description=template.description,
            fact_table=template.fact_table,
            column_count=len(template.columns),
            required_column_count=required_count,
        )

    def _to_column_schema(self, col) -> TemplateColumnSchema:
        return TemplateColumnSchema(
            key=col.key,
            label=col.label,
            required=col.required,
            type=col.column_type,
            description=col.description,
            enum_values=list(col.enum_values),
            example=col.example,
        )


template_service = TemplateService()
